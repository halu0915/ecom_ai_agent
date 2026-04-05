from __future__ import annotations
import sys
from pathlib import Path
import streamlit as st
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from db import get_conn
from query_engine import run_query, build_insights
from reports import query_result_markdown, write_markdown

DB_PATH = ROOT / 'output' / 'db' / 'products.db'

st.set_page_config(page_title='電商資料 AI Agent', layout='wide')
st.title('電商資料 AI Agent')

if not DB_PATH.exists():
    st.warning('尚未建立資料庫，請先執行 scripts/run_etl.py')
    st.stop()

conn = get_conn(str(DB_PATH))

tab1, tab2, tab3 = st.tabs(['儀表板', 'AI 查詢', '即時線上抓取 (Live Fetch)'])

with tab1:
    total = conn.execute('SELECT COUNT(*) FROM products').fetchone()[0]
    brands = conn.execute('SELECT COUNT(DISTINCT brand) FROM products').fetchone()[0]
    cats = conn.execute('SELECT COUNT(DISTINCT category_std) FROM products').fetchone()[0]
    avg_price = conn.execute('SELECT ROUND(AVG(price), 2) FROM products').fetchone()[0]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric('商品數', total)
    c2.metric('品牌數', brands)
    c3.metric('類別數', cats)
    c4.metric('平均價格', avg_price)

    st.subheader('熱門類別')
    top_cats = pd.read_sql_query(
        'SELECT category_std, COUNT(*) AS cnt, ROUND(AVG(price),2) AS avg_price, MAX(sold_30d) AS max_sold FROM products GROUP BY category_std ORDER BY cnt DESC', conn
    )
    st.dataframe(top_cats, use_container_width=True)

with tab2:
    st.subheader('自然語言查詢')
    user_query = st.text_input('輸入查詢', value='找跳蛋類，價格 1000 以下，銷量前 10')
    if st.button('查詢'):
        df, filters, sql = run_query(conn, user_query)
        insights = build_insights(df)
        st.write('解析條件：', filters)
        st.code(sql)
        st.dataframe(
            df, 
            use_container_width=True, 
            column_config={
                "source_url": st.column_config.LinkColumn("商品網址")
            }
        )
        st.markdown('### AI 洞察')
        for item in insights:
            st.write(f'- {item}')

        md = query_result_markdown(user_query, df, insights)
        report_path = ROOT / 'reports' / 'query' / 'latest_query_result.md'
        write_markdown(report_path, md)
        st.success(f'已輸出 Markdown：{report_path}')

with tab3:
    st.subheader('即時線上同步資料引擎')
    st.markdown('不必依賴外部工具匯出 CSV，直接從介面去真實網站爬取資料（此處以開放 API 較友善的 **PChome** 作為首發示範）。')
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        target_platform = st.selectbox('選擇欲即時追蹤的電商', ['PChome', 'Shopee (透過 Playwright)', '中華黃頁 (IYP) - 企業名錄', '台灣黃頁 (Web66)'])
    with col2:
        fetch_keyword = st.text_input('輸入想抓取的商品關鍵字', value='五金')
    with col3:
        scraper_city = 'all'
        if target_platform in ['中華黃頁 (IYP) - 企業名錄', '台灣黃頁 (Web66)']:
            scraper_city = st.selectbox(
                '選擇特定區域', 
                options=['all', 'Taipei', 'NewTaipei', 'Taoyuan', 'Taichung', 'Tainan', 'Kaohsiung'],
                format_func=lambda x: {'all':'全區', 'Taipei':'台北', 'NewTaipei':'新北', 'Taoyuan':'桃園', 'Taichung':'台中', 'Tainan':'台南', 'Kaohsiung':'高雄'}.get(x, x)
            )
            
    if st.button('🚀 立刻去抓即時真實資料！'):
        if target_platform == 'PChome':
            import requests
            url = f"https://ecshweb.pchome.com.tw/search/v3.3/all/results?q={fetch_keyword}&page=1&sort=sale/dc"
            with st.spinner('連線真實 PChome 伺服器中...'):
                try:
                    res = requests.get(url, timeout=10)
                    data = res.json()
                    if 'prods' in data and len(data['prods']) > 0:
                        prods = data['prods']
                        df_live = pd.DataFrame([{
                            '平台': 'PChome',
                            '商品名稱': p.get('name'),
                            '價格': p.get('price'),
                            '真實頁面網址': f"https://24h.pchome.com.tw/prod/{p.get('Id')}"
                        } for p in prods])
                        st.success(f"成功連線！已從 PChome 抓回 {len(prods)} 筆最新商品資料。")
                        st.session_state['live_df'] = df_live
                    else:
                        st.warning("該平台找不到此商品。")
                except Exception as e:
                    st.error(f"即時連線失敗，請稍後再試：{e}")
        elif target_platform == '中華黃頁 (IYP) - 企業名錄':
            import subprocess
            import json
            with st.spinner('連線中華黃頁查詢企業清單並進行二級深度爬取中 (可能需 30~50 秒)...'):
                try:
                    result = subprocess.run(
                        [sys.executable, 'scripts/iyp_scraper.py', fetch_keyword, scraper_city],
                        capture_output=True,
                        text=True,
                        cwd=str(ROOT)
                    )
                    out = result.stdout.strip()
                    if not out:
                        st.error(f"執行失敗: {result.stderr}")
                    else:
                        try:
                            data = json.loads(out)
                        except json.JSONDecodeError:
                            st.error(f"JSON 解析失敗，原始輸出: {out}")
                            data = {}
                        if data.get('success') and data.get('data'):
                            prods = data['data']
                            df_live = pd.DataFrame([{
                                '平台': '中華黃頁 IYP',
                                '名稱': p.get('名稱'),
                                '電話': p.get('電話'),
                                '傳真': p.get('傳真'),
                                '信箱': p.get('信箱'),
                                '地址': p.get('地址'),
                                '真實頁面網址': p.get('真實頁面網址')
                            } for p in prods])
                            st.success(f"成功連線！已抓回 {len(prods)} 筆企業資料。")
                            st.session_state['live_df'] = df_live
                        else:
                            st.error(f"抓取無結果: {data.get('error', '未知錯誤')}")
                except Exception as e:
                    st.error(f"系統錯誤: {e}")
        elif target_platform == '台灣黃頁 (Web66)':
            import subprocess
            import json
            with st.spinner('連線台灣黃頁查詢企業清單並進行二級深度爬取中 (可能需 30~50 秒)...'):
                try:
                    result = subprocess.run(
                        [sys.executable, 'scripts/web66_scraper.py', fetch_keyword, scraper_city],
                        capture_output=True,
                        text=True,
                        cwd=str(ROOT)
                    )
                    out = result.stdout.strip()
                    if not out:
                        st.error(f"執行失敗: {result.stderr}")
                    else:
                        try:
                            data = json.loads(out)
                        except json.JSONDecodeError:
                            st.error(f"JSON 解析失敗，原始輸出: {out}")
                            data = {}
                        if data.get('success') and data.get('data'):
                            prods = data['data']
                            df_live = pd.DataFrame([{
                                '平台': '台灣黃頁 Web66',
                                '名稱': p.get('名稱'),
                                '電話': p.get('電話'),
                                '傳真': p.get('傳真'),
                                '信箱': p.get('信箱'),
                                '地址': p.get('地址'),
                                '真實頁面網址': p.get('真實頁面網址')
                            } for p in prods])
                            st.success(f"成功連線！已抓回 {len(prods)} 筆企業資料。")
                            st.session_state['live_df'] = df_live
                        else:
                            st.error(f"抓取無結果: {data.get('error', '未知錯誤')}")
                except Exception as e:
                    st.error(f"系統錯誤: {e}")
        elif target_platform == 'Shopee (透過 Playwright)':
            import subprocess
            import json
            with st.spinner('驅動 Playwright 無頭瀏覽器潛入 Shopee 中 (約需 10~20 秒，請耐心等候)...'):
                try:
                    result = subprocess.run(
                        [sys.executable, 'scripts/shopee_scraper.py', fetch_keyword],
                        capture_output=True,
                        text=True,
                        cwd=str(ROOT)
                    )
                    out = result.stdout.strip()
                    if not out:
                        st.error(f"Playwright 執行失敗: {result.stderr}")
                    else:
                        try:
                            data = json.loads(out)
                        except json.JSONDecodeError:
                            st.error(f"JSON 解析失敗，原始輸出: {out}")
                            data = {}
                        if data.get('success') and data.get('data'):
                            prods = data['data']
                            df_live = pd.DataFrame([{
                                '平台': 'Shopee',
                                '商品名稱': p.get('名稱'),
                                '價格': p.get('價格'),
                                '真實頁面網址': p.get('網址')
                            } for p in prods])
                            st.success(f"成功連線！已透過 Playwright 抓回 {len(prods)} 筆蝦皮最新商品。")
                            st.session_state['live_df'] = df_live
                        else:
                            st.error(f"抓取無結果或被阻擋: {data.get('error', '未知錯誤')}")
                except Exception as e:
                    st.error(f"系統錯誤: {e}")
                    
    if 'live_df' in st.session_state and not st.session_state['live_df'].empty:
        df_show = st.session_state['live_df']
        st.markdown('---')
        st.subheader('抓取結果')
        
        # Determine column config generically
        col_cfg = {}
        if '真實頁面網址' in df_show.columns:
            col_cfg["真實頁面網址"] = st.column_config.LinkColumn("真實頁面網址")
            
        st.dataframe(df_show, use_container_width=True, column_config=col_cfg)
        
        # Excel Download with Styling
        import io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        def get_styled_excel(df):
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='企業名錄')
                workbook = writer.book
                worksheet = workbook['企業名錄']
                
                # Styles
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=12)
                alt_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                border = Border(left=Side(style='thin', color='BFBFBF'),
                                right=Side(style='thin', color='BFBFBF'),
                                top=Side(style='thin', color='BFBFBF'),
                                bottom=Side(style='thin', color='BFBFBF'))
                center_align = Alignment(horizontal='center', vertical='center')
                left_wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Column widths mapping
                col_widths = {
                    '平台': 15, '名稱': 30, '商品名稱': 40, '電話': 18, '傳真': 18,
                    '信箱': 30, '地址': 45, '描述': 60, '真實頁面網址': 40, '價格': 12
                }
                
                # Apply headers
                for i, col in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=i)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                    
                    # Set width
                    col_letter = cell.column_letter
                    worksheet.column_dimensions[col_letter].width = col_widths.get(col, 20)
                    
                    # Apply row styles
                    for row in range(2, len(df) + 2):
                        c = worksheet.cell(row=row, column=i)
                        c.border = border
                        if row % 2 == 0:
                            c.fill = alt_fill
                        
                        if col in ['描述', '地址', '真實頁面網址', '名稱', '商品名稱']:
                            c.alignment = left_wrap_align
                        else:
                            c.alignment = center_align
            return buf.getvalue()

        excel_data = get_styled_excel(df_show)
            
        dl_col1, dl_col2, dl_col3 = st.columns([2, 2, 2])
        with dl_col1:
            st.download_button(
                label="📥 下載編排版 Excel",
                data=excel_data,
                file_name=f"live_{target_platform}_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with dl_col2:
            st.download_button(
                label="📝 下載 Markdown 檔案",
                data=df_show.to_markdown(index=False),
                file_name="live_fetch_results.md",
                mime="text/markdown",
                use_container_width=True
            )
        with dl_col3:
            if st.button("💾 儲存至專案目錄", use_container_width=True):
                save_dir = ROOT / 'output' / 'live_fetch'
                save_dir.mkdir(parents=True, exist_ok=True)
                file_path = save_dir / f"styled_results_{target_platform}.xlsx"
                with open(file_path, 'wb') as f:
                    f.write(excel_data)
                st.toast(f"已儲存高畫質 Excel 至：{file_path}")
